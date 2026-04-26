"""Generate four test-scenario xlsx fixtures by mutating the supplied gold workbook.

Run with: python3 scripts/build_scenarios.py

Scenarios:
  1. clean_match           — every recon row ties; composite bridges sum to 0
  2. amount_diffs          — clean baseline + specific Internal mutations producing known diffs
  3. new_stripe_rc         — clean baseline + an unmapped Stripe rc not in the mapping
  4. new_internal_rc       — clean baseline + an unmapped Internal rc not in the mapping

The four fixtures are written to fixtures/scenarios/. Tests in tests/test_scenarios.py
assert the expected behaviour for each.
"""
from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
GOLD = ROOT / "fixtures" / "threshold_recon_input.xlsx"
OUT_DIR = ROOT / "fixtures" / "scenarios"

STRIPE_SHEET = "Data - Stripe"
INTERNAL_SHEET = "Data - Internal"

# Per-scenario period labels (different so manifest entries don't collide).
SCENARIO_PERIODS = {
    "scenario_1_clean_match": datetime(2026, 1, 31),
    "scenario_2_amount_diffs": datetime(2026, 2, 28),
    "scenario_3_new_stripe_rc": datetime(2026, 3, 31),
    "scenario_4_new_internal_rc": datetime(2026, 4, 30),
}


def _clone_gold() -> openpyxl.Workbook:
    """Return a fresh load of the gold workbook with only the data sheets retained."""
    wb = openpyxl.load_workbook(GOLD)
    for s in list(wb.sheetnames):
        if s not in (STRIPE_SHEET, INTERNAL_SHEET):
            del wb[s]
    return wb


# --- Stripe sheet helpers ----------------------------------------------------

def _stripe_row_index(ws, rc: str) -> int | None:
    """Return the 1-based row index for the named rc on the Stripe sheet."""
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == rc:
            return r
    return None


def _set_stripe_cells(ws, rc: str, *, gross=None, fee=None) -> None:
    r = _stripe_row_index(ws, rc)
    if r is None:
        raise KeyError(f"Stripe rc '{rc}' not found")
    if gross is not None:
        ws.cell(r, 4, gross)
    if fee is not None:
        ws.cell(r, 5, fee)
    # Recompute NET = gross + fee
    g = ws.cell(r, 4).value or 0
    f = ws.cell(r, 5).value or 0
    ws.cell(r, 6, g + f)


def _zero_stripe_rc(ws, rc: str) -> None:
    _set_stripe_cells(ws, rc, gross=0, fee=0)


def _append_stripe_row(ws, rc: str, gross: float, fee: float = 0) -> int:
    """Insert a new Stripe rc row before the totals footer."""
    # Find the 'total' footer row
    for r in range(4, ws.max_row + 2):
        if ws.cell(r, 1).value is None or str(ws.cell(r, 1).value).lower() == "total":
            insert_at = r
            break
    ws.insert_rows(insert_at)
    ws.cell(insert_at, 1, rc)
    ws.cell(insert_at, 2, "usd")
    ws.cell(insert_at, 3, 1)
    ws.cell(insert_at, 4, gross)
    ws.cell(insert_at, 5, fee)
    ws.cell(insert_at, 6, gross + fee)
    return insert_at


# --- Internal sheet helpers --------------------------------------------------

def _internal_row_index(ws, transaction_category: str, reporting_category: str) -> int | None:
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == transaction_category and ws.cell(r, 2).value == reporting_category:
            return r
    return None


def _set_internal_amount(ws, transaction_category: str, reporting_category: str, amount: float) -> None:
    r = _internal_row_index(ws, transaction_category, reporting_category)
    if r is None:
        raise KeyError(f"Internal row ({transaction_category}, {reporting_category}) not found")
    ws.cell(r, 4, amount)


def _bump_internal_amount(ws, transaction_category: str, reporting_category: str, delta: float) -> None:
    r = _internal_row_index(ws, transaction_category, reporting_category)
    if r is None:
        raise KeyError(f"Internal row ({transaction_category}, {reporting_category}) not found")
    current = ws.cell(r, 4).value or 0
    ws.cell(r, 4, current + delta)


def _delete_internal_rows_for_cat(ws, transaction_category: str) -> None:
    """Delete every base-data row for the given transaction_category."""
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == transaction_category:
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)


def _append_internal_row(ws, transaction_category: str, reporting_category: str, amount: float) -> None:
    """Append a new Internal row at the bottom (before any rollup blocks)."""
    insert_at = ws.max_row + 1
    # Find the first rollup row (e.g. total_subscription) and insert before it.
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.startswith(("total_", "subscription_cycle_")):
            insert_at = r
            break
    ws.insert_rows(insert_at)
    ws.cell(insert_at, 1, transaction_category)
    ws.cell(insert_at, 2, reporting_category)
    ws.cell(insert_at, 3, transaction_category + reporting_category)
    ws.cell(insert_at, 4, amount)


def _set_period(ws, dt: datetime) -> None:
    ws.cell(1, 4, dt)


# --- Mutations ---------------------------------------------------------------

def make_clean_match() -> openpyxl.Workbook:
    """Scenario 1: every B-flag row ties; composite bridges net to 0; no Stripe/Internal-only items."""
    wb = _clone_gold()
    s, i = wb[STRIPE_SHEET], wb[INTERNAL_SHEET]
    _set_period(i, SCENARIO_PERIODS["scenario_1_clean_match"])

    # 1) Remove the per-rc Stripe fee on `charge` so its NET = gross → matches Internal
    #    billing/charge sum (which records gross only). The per-rc fee on `fee` (fee_fee
    #    of -$23,884) is left in place since it's already netted into Stripe `fee` NET
    #    and Internal `fee` sum will tie to it once the sign anomaly is corrected below.
    _set_stripe_cells(s, "charge", fee=0)

    # 2) Fix the SIGN ANOMALY on billing/fee. With +576,119 zeroed out, the remaining
    #    Internal fee cells sum to exactly Stripe NET on the fee row.
    _set_internal_amount(i, "billing", "fee", 0)

    # 3) Remove Stripe-only true reconciling items.
    for rc in ("revenue_share", "other_adjustment",
               "payout_minimum_balance_hold", "payout_minimum_balance_release"):
        _zero_stripe_rc(s, rc)

    # 4) Remove Internal-only subscription_tax rows.
    _delete_internal_rows_for_cat(i, "subscription_tax")

    # The pre-existing $0.46 sub-cent rounding diff on network_cost remains and is
    # still treated as tied within the $1.00 tolerance.
    return wb


def make_amount_diffs() -> openpyxl.Workbook:
    """Scenario 2: clean baseline + specific Internal mutations producing known variances."""
    wb = make_clean_match()
    i = wb[INTERNAL_SHEET]
    _set_period(i, SCENARIO_PERIODS["scenario_2_amount_diffs"])

    # Bump billing/refund by +5,000 (Internal too high → row 1.2 variance becomes +5,000).
    _bump_internal_amount(i, "billing", "refund", 5000)
    # Reduce cx_transfer/transfer by 1,500 (Internal too low → row 3.1 variance becomes -1,500).
    _bump_internal_amount(i, "cx_transfer", "transfer", -1500)
    return wb


def make_new_stripe_rc() -> openpyxl.Workbook:
    """Scenario 3: clean baseline + a new Stripe rc not in the mapping."""
    wb = make_clean_match()
    s, i = wb[STRIPE_SHEET], wb[INTERNAL_SHEET]
    _set_period(i, SCENARIO_PERIODS["scenario_3_new_stripe_rc"])

    _append_stripe_row(s, "beta_program_credit", gross=2000, fee=0)
    return wb


def make_new_internal_rc() -> openpyxl.Workbook:
    """Scenario 4: clean baseline + a new Internal rc not in the mapping."""
    wb = make_clean_match()
    i = wb[INTERNAL_SHEET]
    _set_period(i, SCENARIO_PERIODS["scenario_4_new_internal_rc"])

    _append_internal_row(i, "other", "loyalty_credit", 3000)
    return wb


# --- Driver ------------------------------------------------------------------

def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    builders = [
        ("scenario_1_clean_match.xlsx", make_clean_match),
        ("scenario_2_amount_diffs.xlsx", make_amount_diffs),
        ("scenario_3_new_stripe_rc.xlsx", make_new_stripe_rc),
        ("scenario_4_new_internal_rc.xlsx", make_new_internal_rc),
    ]
    for filename, fn in builders:
        wb = fn()
        out = OUT_DIR / filename
        wb.save(out)
        print(f"  → {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
