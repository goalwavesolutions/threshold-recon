"""Read the monthly Stripe + Internal source files into normalised DataFrames.

Stripe and Internal data are uploaded as two separate files (xlsx or csv). Each
reader detects the header row by scanning for expected token names, so files with
title rows (the original Excel layout) and bare CSV exports both work. For xlsx
sources, the named sheet is preferred but the first sheet is used as a fallback.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import IO, Any

import openpyxl
import pandas as pd

from threshold.config import Mapping

STRIPE_SHEET = "Data - Stripe"
INTERNAL_SHEET = "Data - Internal"

# Tokens that uniquely identify the header row in each file.
STRIPE_HEADER_TOKENS = ("reporting_category", "gross", "fee", "net")
INTERNAL_HEADER_TOKENS = ("transaction_category", "reporting_category")

HEADER_SCAN_LIMIT = 20


@dataclass
class SourceData:
    stripe: pd.DataFrame
    internal: pd.DataFrame
    period_label: str
    dropped_rollup_rows: int


# --- raw row loading ---------------------------------------------------------

def _is_csv(filename: str | None) -> bool:
    return bool(filename) and Path(filename).suffix.lower() == ".csv"


def _load_xlsx_rows(source: str | Path | bytes | IO, preferred_sheet: str) -> list[list[Any]]:
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        ws = wb[preferred_sheet] if preferred_sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _load_csv_rows(source: str | Path | bytes | IO) -> list[list[Any]]:
    """Load a CSV into list-of-lists. Auto-sniffs delimiter, treats blanks as None."""
    if isinstance(source, bytes):
        text = source.decode("utf-8-sig", errors="replace")
    elif isinstance(source, (str, Path)):
        with open(source, "rb") as f:
            text = f.read().decode("utf-8-sig", errors="replace")
    else:
        text = source.read()
        if isinstance(text, bytes):
            text = text.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    out = []
    for row in reader:
        out.append([(cell if cell != "" else None) for cell in row])
    return out


def _load_rows(
    source: str | Path | bytes | IO,
    filename: str | None,
    preferred_sheet: str,
) -> list[list[Any]]:
    if _is_csv(filename):
        return _load_csv_rows(source)
    return _load_xlsx_rows(source, preferred_sheet)


# --- header detection --------------------------------------------------------

def _row_text_set(row: list[Any]) -> set[str]:
    return {str(c).strip().lower() for c in row if c is not None}


def _find_header_row(rows: list[list[Any]], required_tokens: tuple[str, ...]) -> int:
    required = {t.lower() for t in required_tokens}
    for i, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        if required.issubset(_row_text_set(row)):
            return i
    raise ValueError(
        f"Could not locate header row containing tokens {required_tokens} "
        f"within the first {HEADER_SCAN_LIMIT} rows."
    )


# --- Stripe + Internal readers ----------------------------------------------

def read_stripe(source: str | Path | bytes | IO, filename: str | None = None) -> pd.DataFrame:
    rows = _load_rows(source, filename, STRIPE_SHEET)
    header_idx = _find_header_row(rows, STRIPE_HEADER_TOKENS)
    header = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(rows[header_idx])]
    data = [r[: len(header)] for r in rows[header_idx + 1:] if any(c is not None for c in r)]
    df = pd.DataFrame(data, columns=header)
    df = df.dropna(subset=["reporting_category"]).copy()
    df = df[df["reporting_category"].astype(str).str.lower() != "total"].reset_index(drop=True)
    for col in ("count", "gross", "fee", "net"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def _parse_period_label(value: Any) -> str:
    if value is None:
        return "unknown"
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    s = str(value).strip()
    # Strip a trailing time component if present (Excel-to-CSV often emits "2025-12-31 00:00:00").
    s_date = s.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m", "%b-%y", "%B %Y"):
        try:
            return datetime.strptime(s_date, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return s or "unknown"


def read_internal(
    source: str | Path | bytes | IO, mapping: Mapping, filename: str | None = None,
) -> tuple[pd.DataFrame, str, int]:
    rows = _load_rows(source, filename, INTERNAL_SHEET)
    header_idx = _find_header_row(rows, INTERNAL_HEADER_TOKENS)
    header_row = rows[header_idx]
    period_label = _parse_period_label(header_row[3] if len(header_row) > 3 else None)

    data = [r[:4] for r in rows[header_idx + 1:] if any(c is not None for c in r)]
    # Pad short rows so the DataFrame constructor doesn't choke on ragged input.
    data = [list(r) + [None] * (4 - len(r)) if len(r) < 4 else r for r in data]
    df = pd.DataFrame(
        data,
        columns=["transaction_category", "reporting_category", "concat_key", "amount"],
    )
    df = df.dropna(subset=["transaction_category", "reporting_category"]).copy()

    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df.dropna(subset=["amount"])

    before = len(df)
    df = df[~df["transaction_category"].isin(mapping.rollup_categories_excluded)].reset_index(drop=True)
    dropped = before - len(df)
    return df, period_label, dropped


# --- orchestrator ------------------------------------------------------------

def load_source(
    stripe_source: str | Path | bytes | IO,
    internal_source: str | Path | bytes | IO,
    mapping: Mapping,
    stripe_filename: str | None = None,
    internal_filename: str | None = None,
) -> SourceData:
    """Load the two source files into a unified SourceData.

    `stripe_filename` and `internal_filename` are used to detect xlsx vs csv. When
    omitted (e.g. the test fixture passes raw bytes), xlsx is assumed. Pass the same
    combined xlsx for both arguments to use the legacy single-file layout — each
    reader will pick its own sheet by name.
    """
    stripe = read_stripe(stripe_source, stripe_filename)
    internal, period, dropped = read_internal(internal_source, mapping, internal_filename)
    return SourceData(stripe=stripe, internal=internal, period_label=period, dropped_rollup_rows=dropped)
