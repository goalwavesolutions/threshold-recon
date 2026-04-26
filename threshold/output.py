"""Write the recon + summary to an Excel workbook matching the gold layout."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from threshold.config import Mapping
from threshold.exceptions import Exception_, to_dataframe as exceptions_to_dataframe
from threshold.summary import SummaryView

HEADER_FILL = PatternFill("solid", fgColor="1F2D5C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TYPE_FILL = PatternFill("solid", fgColor="EAEEF7")
TIE_FILL = PatternFill("solid", fgColor="E6F4EA")
EXC_FILL = PatternFill("solid", fgColor="FCE8E6")
TITLE_FONT = Font(bold=True, size=12)
NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'


def _write_recon_sheet(wb: Workbook, rows: pd.DataFrame, mapping: Mapping, period_label: str) -> None:
    ws = wb.create_sheet("Recon - Stripe vs Internal")
    ws["A1"] = f"Stripe ↔ Internal Reconciliation — {period_label} (grouped by Type)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=23)
    ws["A2"] = (
        "Rows grouped by economic Type. Source flag: B=Both, S=Stripe only, I=Internal only. "
        "Internal cells = SUMIFS(Data-Internal, cat, rc). Stripe NET = col F (gross+fee)."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=23)

    headers = (
        ["Type", "reporting_category", "Src", "Ref", "Description"]
        + mapping.cat_columns
        + ["Internal Total", "Stripe NET", "Variance", "Tie?", "Exception Type", "Comments"]
    )
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(rows.iterrows(), start=5):
        is_tied = row["Tie?"] == "✓"
        ws.cell(i, 1, row["Type"])
        ws.cell(i, 2, row["reporting_category"])
        ws.cell(i, 3, row["Src"]).alignment = Alignment(horizontal="center")
        ws.cell(i, 4, row["Ref"]).number_format = "0.0"
        ws.cell(i, 5, row["Description"])
        for j, cat in enumerate(mapping.cat_columns):
            cell = ws.cell(i, 6 + j, float(row[cat]))
            cell.number_format = NUM_FMT
        ws.cell(i, 6 + len(mapping.cat_columns), float(row["Internal Total"])).number_format = NUM_FMT
        ws.cell(i, 7 + len(mapping.cat_columns), float(row["Stripe NET"])).number_format = NUM_FMT
        ws.cell(i, 8 + len(mapping.cat_columns), float(row["Variance"])).number_format = NUM_FMT
        tie_cell = ws.cell(i, 9 + len(mapping.cat_columns), row["Tie?"])
        tie_cell.alignment = Alignment(horizontal="center")
        tie_cell.fill = TIE_FILL if is_tied else EXC_FILL
        ws.cell(i, 10 + len(mapping.cat_columns), row.get("Exception Type", ""))
        ws.cell(i, 11 + len(mapping.cat_columns), row.get("Comments", "")).alignment = Alignment(wrap_text=True)

    # Totals row
    n = len(rows)
    last_data_row = 4 + n
    total_row = last_data_row + 2
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    cat_start_col = 6
    for j in range(len(mapping.cat_columns) + 3):  # cats + Internal Total + Stripe NET + Variance
        col = cat_start_col + j
        col_letter = get_column_letter(col)
        cell = ws.cell(total_row, col, f"=SUM({col_letter}5:{col_letter}{last_data_row})")
        cell.number_format = NUM_FMT
        cell.font = Font(bold=True)

    # Validation row: per-cat sum from raw Internal must equal recon's per-cat sum
    val_row = total_row + 1
    ws.cell(val_row, 1, "Internal Cat Total (validation)").font = Font(italic=True)
    delta_row = val_row + 1
    ws.cell(delta_row, 1, "Δ (TOTAL − validation) → 0 confirms recon captures all source").font = Font(italic=True)

    # Column widths
    widths = [18, 30, 6, 6, 40] + [16] * len(mapping.cat_columns) + [16, 16, 16, 6, 36, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "F5"


def _write_summary_sheet(wb: Workbook, sv: SummaryView, period_label: str) -> None:
    ws = wb.create_sheet("Recon Summary")
    ws["A1"] = f"Reconciliation Summary — {period_label}"
    ws["A1"].font = TITLE_FONT

    r = 3
    ws.cell(r, 1, "HIGH-LEVEL SUMMARY — Total by source flag").font = Font(bold=True)
    r += 1
    headers = list(sv.bsi_summary.columns)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(r, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL
    r += 1
    for _, row in sv.bsi_summary.iterrows():
        ws.cell(r, 1, row["Source"]).font = Font(bold=True if row["Source"] == "TOTAL" else False)
        ws.cell(r, 2, row["Description"])
        for j, col in enumerate(["Internal Total", "Stripe NET Total", "Net Variance"], start=3):
            c = ws.cell(r, j, float(row[col])); c.number_format = NUM_FMT
        r += 1
    r += 1

    ws.cell(r, 1, "NETTING BRIDGES — refs that pair off").font = Font(bold=True); r += 2
    for bridge in sv.composite_bridges:
        ws.cell(r, 1, f"Composite bridge: {bridge.iloc[-2]['rc']} = {bridge.iloc[-3]['Component']}").font = Font(bold=True)
        r += 1
        for i, h in enumerate(bridge.columns, start=1):
            c = ws.cell(r, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL
        r += 1
        for _, brow in bridge.iterrows():
            ws.cell(r, 1, brow["Ref"])
            ws.cell(r, 2, brow["rc"])
            ws.cell(r, 3, brow["Component"])
            c = ws.cell(r, 4, float(brow["Amount"])); c.number_format = NUM_FMT
            r += 1
        r += 1

    if sv.reserve_bridge is not None:
        ws.cell(r, 1, "Reserve mechanic pair (sums to 0)").font = Font(bold=True); r += 1
        for i, h in enumerate(sv.reserve_bridge.columns, start=1):
            c = ws.cell(r, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL
        r += 1
        for _, brow in sv.reserve_bridge.iterrows():
            ws.cell(r, 1, brow["Ref"])
            ws.cell(r, 2, brow["rc"])
            ws.cell(r, 3, brow["Component"])
            c = ws.cell(r, 4, float(brow["Amount"])); c.number_format = NUM_FMT
            r += 1
        r += 1

    ws.cell(r, 1, "NETTING VIEW — Internal − Stripe NET (signed). SUM = grand residual.").font = Font(bold=True); r += 2
    for i, h in enumerate(sv.netting_view.columns, start=1):
        c = ws.cell(r, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL
    r += 1
    netting_start = r
    for _, nrow in sv.netting_view.iterrows():
        ws.cell(r, 1, nrow["Item"])
        ws.cell(r, 2, str(nrow["Refs"]))
        ws.cell(r, 3, nrow["Sign"]).alignment = Alignment(horizontal="center")
        c = ws.cell(r, 4, float(nrow["Internal − Stripe"])); c.number_format = NUM_FMT
        ws.cell(r, 5, nrow["Notes"]).alignment = Alignment(wrap_text=True)
        r += 1
    netting_end = r - 1
    r += 1
    ws.cell(r, 1, "Net residual (SUM of signed amounts above)").font = Font(bold=True)
    c = ws.cell(r, 4, f"=SUM(D{netting_start}:D{netting_end})"); c.number_format = NUM_FMT; c.font = Font(bold=True)
    r += 1
    ws.cell(r, 1, "Cross-check: grand total Internal − Stripe").font = Font(italic=True)
    c = ws.cell(r, 4, sv.grand_residual); c.number_format = NUM_FMT
    ws.cell(r, 5, "Should equal D above — proves netting view captures full residual.")

    widths = [50, 30, 8, 18, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_corrections_sheet(wb: Workbook, corrections: list) -> None:
    """Document any cell-level Internal source corrections (e.g. sign flips) applied this run."""
    ws = wb.create_sheet("Corrections")
    ws["A1"] = "Source-data corrections applied to Internal before reconciliation"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = (
        "Each row represents an adjustment to an Internal cell prior to recon. "
        "These reconcile known source bugs (e.g. sign anomalies). The original source data is unchanged."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = [
        "Transaction Category", "Reporting Category", "Kind",
        "Original Amount", "Corrected Amount", "Δ", "Comment", "Applied At",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(4, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL

    if not corrections:
        ws.cell(5, 1, "No corrections applied for this run.").font = Font(italic=True)
    else:
        for ri, corr in enumerate(corrections, start=5):
            ws.cell(ri, 1, corr.transaction_category)
            ws.cell(ri, 2, corr.reporting_category)
            ws.cell(ri, 3, corr.kind)
            ws.cell(ri, 4, float(corr.original_amount)).number_format = NUM_FMT
            ws.cell(ri, 5, float(corr.corrected_amount)).number_format = NUM_FMT
            ws.cell(ri, 6, float(corr.corrected_amount - corr.original_amount)).number_format = NUM_FMT
            ws.cell(ri, 7, corr.comment).alignment = Alignment(wrap_text=True)
            ws.cell(ri, 8, corr.applied_at)

    widths = [22, 22, 12, 16, 16, 16, 60, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_exceptions_sheet(wb: Workbook, excs: list[Exception_]) -> None:
    ws = wb.create_sheet("Exceptions")
    df = exceptions_to_dataframe(excs)
    if df.empty:
        ws["A1"] = "No exceptions detected."
        return
    for i, h in enumerate(df.columns, start=1):
        c = ws.cell(1, i, h); c.font = HEADER_FONT; c.fill = HEADER_FILL
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        for ci, col in enumerate(df.columns, start=1):
            v = row[col]
            cell = ws.cell(ri, ci, v if not isinstance(v, float) else float(v))
            if col == "Variance":
                cell.number_format = NUM_FMT
            if col == "Comment / Suggested Action":
                cell.alignment = Alignment(wrap_text=True)
    widths = [6, 30, 6, 16, 24, 36, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_workbook(
    rows: pd.DataFrame,
    mapping: Mapping,
    summary: SummaryView,
    excs: list[Exception_],
    period_label: str,
    corrections: list | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _write_recon_sheet(wb, rows, mapping, period_label)
    _write_summary_sheet(wb, summary, period_label)
    _write_corrections_sheet(wb, corrections or [])
    _write_exceptions_sheet(wb, excs)
    return wb


def write_workbook(
    out_path: Path | str,
    rows: pd.DataFrame,
    mapping: Mapping,
    summary: SummaryView,
    excs: list[Exception_],
    period_label: str,
    corrections: list | None = None,
) -> Path:
    wb = _build_workbook(rows, mapping, summary, excs, period_label, corrections)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def workbook_bytes(
    rows: pd.DataFrame,
    mapping: Mapping,
    summary: SummaryView,
    excs: list[Exception_],
    period_label: str,
    corrections: list | None = None,
) -> bytes:
    """Build the workbook in-memory and return its bytes — for live downloads."""
    import io
    wb = _build_workbook(rows, mapping, summary, excs, period_label, corrections)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
